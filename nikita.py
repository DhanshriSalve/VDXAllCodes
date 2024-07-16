from openpyxl import Workbook
import time

wb = Workbook()
sheet = wb.active

sheet['A1'] = 87
sheet['A2'] = "Dhanshri"
sheet['A3'] = 41.80
sheet['A4'] = 10

sheet1['A10']=40

# Corrected create_sheet methods
sheet1 = wb.create_sheet("sheet1", 0)
sheet2 = wb.create_sheet("sheet2")
sheet3 = wb.create_sheet("sheet3")

# Using current date as value for A5
n = time.strftime("%x")
sheet['A5'] = n

# Setting value in cell B2
sheet.cell(row=2, column=2).value = 5

# Renaming sheet
sheet.title = "dhanu"

# Saving the workbook
wb.save("nikita.xlsx")
